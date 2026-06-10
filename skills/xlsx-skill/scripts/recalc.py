"""Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice.
Requires LibreOffice to be installed.

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_DIR_WINDOWS = os.path.expandvars(
    r"%APPDATA%\LibreOffice\4\user\basic\Standard"
)
MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def setup_libreoffice_macro():
    system = platform.system()
    if system == "Darwin":
        macro_dir = os.path.expanduser(MACRO_DIR_MACOS)
    elif system == "Windows":
        macro_dir = MACRO_DIR_WINDOWS
    else:
        macro_dir = os.path.expanduser(MACRO_DIR_LINUX)

    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        soffice_cmd = "soffice" if system != "Windows" else "soffice.exe"
        subprocess.run(
            [soffice_cmd, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File not found: {filename}"}

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    abs_path = str(Path(filename).resolve())
    soffice_cmd = "soffice" if platform.system() != "Windows" else "soffice.exe"

    cmd = [
        soffice_cmd, "--headless", "--invisible",
        f"macro:///Standard.Module1.RecalculateAndSave",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Windows":
        # Windows doesn't have timeout command for this use
        pass

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout + 5)

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        return {"error": error_msg}

    # Scan for Excel errors
    try:
        wb = load_workbook(filename, data_only=True)
        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        total_formulas = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1

        # Count formulas from non-data-only load
        wb_formulas = load_workbook(filename)
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1

        output = {
            "status": "errors_found" if total_errors > 0 else "success",
            "total_errors": total_errors,
            "total_formulas": total_formulas,
        }

        if total_errors > 0:
            output["error_summary"] = {
                err: {"count": len(locs), "locations": locs}
                for err, locs in error_details.items() if locs
            }

        return output

    except Exception as e:
        return {"error": f"Error scanning file: {str(e)}"}


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))

    if "error" in result or result.get("total_errors", 0) > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
